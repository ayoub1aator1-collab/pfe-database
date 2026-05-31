"""
=============================================================
  ONCF Z2M — Système de Supervision VFD
  Fichier  : logger_7.py
  Rôle     : Backend FastAPI — lecture Arduino / simulation
             + WebSocket, historique, export Excel,
             + contrôle manuel (sliders depuis React)
  Auteur   : PFE 2024-2025
  Version  : 5.0
=============================================================
"""

# ─────────────────────────────────────────
#  IMPORTS
# ─────────────────────────────────────────
import asyncio
import io
import json
import random
import sys
import threading
import time
from collections import deque
from datetime import datetime, timedelta
from pathlib import Path

# Fix Windows encoding (cp1256 ne supporte pas les emojis)
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')

try:
    import serial
    import serial.tools.list_ports
except ImportError:
    serial = None
    print("⚠️  pyserial non installé — mode simulation uniquement")

try:
    import uvicorn
    from fastapi import FastAPI, WebSocket, WebSocketDisconnect, Query
    from fastapi.middleware.cors import CORSMiddleware
    from fastapi.responses import StreamingResponse
    from pydantic import BaseModel
except ImportError:
    print("❌ Erreur: pip install fastapi uvicorn")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    openpyxl = None
    print("⚠️  openpyxl non installé — export Excel désactivé")


# ─────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────
SERIAL_PORT = "COM2"
BAUD_RATE = 9600
POLL_INTERVAL = 1.0
HISTORY_MAXLEN = 3600

# Seuils d'alarme
config = {
    "alarm_temp": 45.0,
    "alarm_pression": 20.0,
    "warn_temp": 42.0,
    "warn_pression": 18.0,
    "force_mode": None,
}
config_lock = threading.Lock()


# ─────────────────────────────────────────
#  ÉTAT GLOBAL
# ─────────────────────────────────────────
latest_data = {
    "time": "--:--:--",
    "temp_in": 0.0,
    "temp_out": 0.0,
    "delta_t": 0.0,
    "pressure": 0.0,
    "vitesse": 0,
    "motors": 0,
    "freq": 10.0,
    "mode": "simulation",
    "alarm": False,
    "warning": False,
    "system_emergency": False,
}
data_lock = threading.Lock()

# Historique en mémoire
history = deque(maxlen=HISTORY_MAXLEN)

# Mode manuel (contrôle depuis React)
manual_mode = False
manual_data = {
    "temp_in": 25.0,
    "temp_out": 35.0,
    "pressure": 15.0,
    "vitesse": 1500,
}
manual_lock = threading.Lock()


# ─────────────────────────────────────────
#  WEBSOCKET — Gestion des connexions
# ─────────────────────────────────────────
class ConnectionManager:
    """Gère les connexions WebSocket actives."""

    def __init__(self):
        self.active_connections = []
        self._lock = asyncio.Lock()

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        async with self._lock:
            self.active_connections.append(websocket)
        print(f"🔌 WebSocket connecté — {len(self.active_connections)} client(s)")

    async def disconnect(self, websocket: WebSocket):
        async with self._lock:
            if websocket in self.active_connections:
                self.active_connections.remove(websocket)
        print(f"🔌 WebSocket déconnecté — {len(self.active_connections)} client(s)")

    async def broadcast(self, data: dict):
        """Envoie les données à tous les clients connectés."""
        async with self._lock:
            dead = []
            for ws in self.active_connections:
                try:
                    await ws.send_json(data)
                except Exception:
                    dead.append(ws)
            for ws in dead:
                self.active_connections.remove(ws)


manager = ConnectionManager()
_loop = None


def _broadcast_data(data: dict):
    """Envoie les données via WebSocket depuis le thread de lecture."""
    global _loop
    if _loop and manager.active_connections:
        asyncio.run_coroutine_threadsafe(manager.broadcast(data), _loop)


# ─────────────────────────────────────────
#  SIMULATION RÉALISTE
# ─────────────────────────────────────────
_sim_temp_in = 25.0
_sim_temp_out = 35.0
_sim_press = 15.0
_sim_direction = 1


def generer_donnees_simulation() -> dict:
    """
    Génère des données de simulation réalistes.
    Logique ventilateurs basée sur la pression :
      0  – 12 bar  →  1 moteur
      12 – 16 bar  →  2 moteurs
      18 – 25 bar  →  3 moteurs
      > 25 bar     →  ⛔ ARRÊT SYSTÈME
    """
    global _sim_temp_in, _sim_temp_out, _sim_press, _sim_direction

    with config_lock:
        alarm_temp = config["alarm_temp"]
        alarm_press = config["alarm_pression"]
        warn_temp = config["warn_temp"]
        warn_press = config["warn_pression"]

    # Inertie thermique entrée (varie peu)
    _sim_temp_in += random.uniform(-0.08, 0.08)
    _sim_temp_in = max(20.0, min(35.0, _sim_temp_in))

    # Inertie thermique sortie (varie plus)
    _sim_temp_out += _sim_direction * 0.12 + random.uniform(-0.06, 0.06)
    if _sim_temp_out >= 48.0:
        _sim_direction = -1
    elif _sim_temp_out <= 28.0:
        _sim_direction = 1
    _sim_temp_out = max(25.0, min(50.0, _sim_temp_out))

    # Pression avec inertie
    _sim_press += random.uniform(-0.15, 0.15)
    _sim_press = max(0.0, min(30.0, _sim_press))

    # Delta T
    delta_t = round(abs(_sim_temp_out - _sim_temp_in), 1)

    # Logique ventilateurs
    system_emergency = False
    if _sim_press > 25.0:
        motors = 3
        system_emergency = True
    elif _sim_press >= 18.0:
        motors = 3
    elif _sim_press >= 12.0:
        motors = 2
    else:
        motors = 1

    # Fréquence et vitesse
    freq = round(max(10.0, min(80.0, (_sim_press / 25.0) * 80)), 1)
    vitesse = int(freq * 30)

    return {
        "time": datetime.now().strftime("%H:%M:%S"),
        "temp_in": round(_sim_temp_in, 1),
        "temp_out": round(_sim_temp_out, 1),
        "delta_t": delta_t,
        "pressure": round(_sim_press, 1),
        "motors": motors,
        "freq": freq,
        "vitesse": vitesse,
        "mode": "simulation",
        "alarm": _sim_temp_out > alarm_temp or _sim_press > alarm_press,
        "warning": _sim_temp_out > warn_temp or _sim_press > warn_press,
        "system_emergency": system_emergency,
    }


# ─────────────────────────────────────────
#  DONNÉES MANUELLES (depuis les sliders React)
# ─────────────────────────────────────────
def generer_donnees_manuelles() -> dict:
    """
    Construit un point de données à partir des valeurs
    envoyées manuellement par l'interface React (sliders).
    """
    with manual_lock:
        t_in = manual_data["temp_in"]
        t_out = manual_data["temp_out"]
        press = manual_data["pressure"]
        vit = manual_data["vitesse"]

    with config_lock:
        alarm_temp = config["alarm_temp"]
        alarm_press = config["alarm_pression"]
        warn_temp = config["warn_temp"]
        warn_press = config["warn_pression"]

    delta_t = round(abs(t_out - t_in), 1)

    # Logique ventilateurs
    system_emergency = False
    if press > 25.0:
        motors = 3
        system_emergency = True
    elif press >= 18.0:
        motors = 3
    elif press >= 12.0:
        motors = 2
    else:
        motors = 1

    # Fréquence calculée depuis la pression
    freq = round(max(10.0, min(80.0, (press / 25.0) * 80)), 1)

    return {
        "time": datetime.now().strftime("%H:%M:%S"),
        "temp_in": round(t_in, 1),
        "temp_out": round(t_out, 1),
        "delta_t": delta_t,
        "pressure": round(press, 1),
        "motors": motors,
        "freq": freq,
        "vitesse": vit,
        "mode": "manuel",
        "alarm": t_out > alarm_temp or press > alarm_press,
        "warning": t_out > warn_temp or press > warn_press,
        "system_emergency": system_emergency,
    }


# ─────────────────────────────────────────
#  PARSEUR JSON ARDUINO
#  Format: {"T_IN":32.1,"T_OUT":38.5,"P":15.2,"F":48}
# ─────────────────────────────────────────
def parser_arduino(ligne: str) -> dict | None:
    """Parse la ligne JSON reçue de l'Arduino."""
    try:
        ligne = ligne.strip()
        if not ligne.startswith("{"):
            return None

        obj = json.loads(ligne)

        with config_lock:
            alarm_temp = config["alarm_temp"]
            alarm_press = config["alarm_pression"]
            warn_temp = config["warn_temp"]
            warn_press = config["warn_pression"]

        t_in = float(obj.get("T_IN", 25.0))
        t_out = float(obj.get("T_OUT", t_in + 5.0))
        press = float(obj.get("P", 0.0))
        freq = float(obj.get("F", 10.0))

        delta_t = round(abs(t_out - t_in), 1)
        vitesse = int(freq * 30)

        # Logique ventilateurs
        system_emergency = False
        if press > 25.0:
            motors = 3
            system_emergency = True
        elif press >= 18.0:
            motors = 3
        elif press >= 12.0:
            motors = 2
        else:
            motors = 1

        return {
            "time": datetime.now().strftime("%H:%M:%S"),
            "temp_in": round(t_in, 1),
            "temp_out": round(t_out, 1),
            "delta_t": delta_t,
            "pressure": round(press, 1),
            "motors": motors,
            "freq": round(freq, 1),
            "vitesse": vitesse,
            "mode": "arduino",
            "alarm": t_out > alarm_temp or press > alarm_press,
            "warning": t_out > warn_temp or press > warn_press,
            "system_emergency": system_emergency,
        }

    except (json.JSONDecodeError, KeyError, ValueError) as e:
        print(f"⚠️  Erreur parsing Arduino: {e} | Ligne: {ligne!r}")
        return None


# ─────────────────────────────────────────
#  THREAD DE LECTURE
# ─────────────────────────────────────────
def boucle_lecture():
    """
    Thread principal :
    - Si mode manuel activé → utilise les sliders React
    - Sinon tente Arduino
    - Sinon fallback simulation
    """
    global latest_data
    ser = None

    print(f"📡 Tentative d'ouverture de {SERIAL_PORT}...")

    while True:
        # ── Mode manuel ? ────────────────────────────
        if manual_mode:
            data = generer_donnees_manuelles()
            with data_lock:
                latest_data = data
            data_with_ts = {**data, "timestamp": datetime.now().isoformat()}
            history.append(data_with_ts)
            _broadcast_data(data)

            emoji = "🔴" if data["alarm"] else ("🟡" if data["warning"] else "🟢")
            print(
                f"🎛️  Manuel [{data['time']}] {emoji} "
                f"In:{data['temp_in']}°C  Out:{data['temp_out']}°C  "
                f"dT:{data['delta_t']}°C  P:{data['pressure']}bar  "
                f"V:{data['vitesse']}RPM"
            )
            time.sleep(POLL_INTERVAL)
            continue

        # ── Vérifier si mode forcé simulation ────────
        with config_lock:
            force_mode = config["force_mode"]

        # ── Tentative connexion Arduino ──────────────
        if force_mode != "simulation" and serial is not None:
            if ser is None or not ser.is_open:
                try:
                    ser = serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=2)
                    ser.flushInput()
                    print(f"✅ Arduino connecté sur {SERIAL_PORT} @ {BAUD_RATE} baud")
                except Exception as e:
                    if ser:
                        try:
                            ser.close()
                        except Exception:
                            pass
                    ser = None

        # ── Lecture Arduino ──────────────────────────
        if force_mode != "simulation" and ser and ser.is_open:
            try:
                if ser.in_waiting > 0:
                    ligne = ser.readline().decode("utf-8", errors="ignore")
                    data = parser_arduino(ligne)
                    if data:
                        with data_lock:
                            latest_data = data
                        data_with_ts = {**data, "timestamp": datetime.now().isoformat()}
                        history.append(data_with_ts)
                        _broadcast_data(data)

                        emoji = "🔴" if data["alarm"] else ("🟡" if data["warning"] else "🟢")
                        print(
                            f"📡 Arduino [{data['time']}] {emoji} "
                            f"In:{data['temp_in']}°C  Out:{data['temp_out']}°C  "
                            f"dT:{data['delta_t']}°C  P:{data['pressure']}bar  "
                            f"V:{data['vitesse']}RPM"
                        )
            except Exception:
                print("❌ Connexion Arduino perdue — basculement simulation")
                try:
                    ser.close()
                except Exception:
                    pass
                ser = None

        # ── Mode simulation (fallback) ───────────────
        else:
            data = generer_donnees_simulation()
            with data_lock:
                latest_data = data
            data_with_ts = {**data, "timestamp": datetime.now().isoformat()}
            history.append(data_with_ts)
            _broadcast_data(data)

            emoji = "🔴" if data["alarm"] else ("🟡" if data["warning"] else "🟢")
            print(
                f"📊 Simul [{data['time']}] {emoji} "
                f"In:{data['temp_in']}°C  Out:{data['temp_out']}°C  "
                f"dT:{data['delta_t']}°C  P:{data['pressure']}bar  "
                f"V:{data['vitesse']}RPM"
            )

        time.sleep(POLL_INTERVAL)


# ─────────────────────────────────────────
#  APPLICATION FASTAPI
# ─────────────────────────────────────────
app = FastAPI(
    title="ONCF Z2M — Supervision VFD",
    description="API temps réel pour le refroidissement du train Z2M",
    version="5.0",
)

# CORS — autoriser React (dev + prod)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup():
    """Capture la boucle asyncio pour le broadcast WebSocket."""
    global _loop
    _loop = asyncio.get_running_loop()


# ── Endpoint : Dernière mesure ────────────────────────
@app.get("/data", summary="Dernière mesure temps réel")
def get_data() -> dict:
    """Retourne la dernière valeur lue (Arduino, simulation, ou manuel)."""
    with data_lock:
        return dict(latest_data)


# ── Endpoint : État du serveur ────────────────────────
@app.get("/health", summary="État du serveur")
def health() -> dict:
    return {
        "status": "ok",
        "source": latest_data.get("mode", "?"),
        "clients_ws": len(manager.active_connections),
        "history_size": len(history),
        "manual_mode": manual_mode,
    }


# ── Endpoint : Historique ─────────────────────────────
@app.get("/history", summary="Historique des mesures")
def get_history(minutes: int = Query(default=30, ge=1, le=60)) -> dict:
    """Retourne les mesures des N dernières minutes."""
    cutoff = datetime.now() - timedelta(minutes=minutes)
    cutoff_iso = cutoff.isoformat()
    filtered = [
        point for point in history
        if point.get("timestamp", "") >= cutoff_iso
    ]
    return {
        "count": len(filtered),
        "minutes": minutes,
        "data": list(filtered),
    }


# ── Endpoint : Statistiques ──────────────────────────
@app.get("/stats", summary="Statistiques min/max/avg")
def get_stats(minutes: int = Query(default=30, ge=1, le=60)) -> dict:
    """Calcule min, max et moyenne sur les N dernières minutes."""
    cutoff = datetime.now() - timedelta(minutes=minutes)
    cutoff_iso = cutoff.isoformat()
    filtered = [
        p for p in history
        if p.get("timestamp", "") >= cutoff_iso
    ]
    if not filtered:
        return {
            "count": 0,
            "minutes": minutes,
        }

    def calc(values):
        return {
            "min": round(min(values), 1),
            "max": round(max(values), 1),
            "avg": round(sum(values) / len(values), 1),
        }

    return {
        "count": len(filtered),
        "minutes": minutes,
        "temp_in": calc([p["temp_in"] for p in filtered]),
        "temp_out": calc([p["temp_out"] for p in filtered]),
        "delta_t": calc([p["delta_t"] for p in filtered]),
        "pressure": calc([p["pressure"] for p in filtered]),
        "freq": calc([p["freq"] for p in filtered]),
        "vitesse": calc([p["vitesse"] for p in filtered]),
        "alarms": sum(1 for p in filtered if p.get("alarm")),
        "warnings": sum(1 for p in filtered if p.get("warning")),
    }


# ── Modèle Pydantic pour le contrôle manuel ──────────
class ManualControl(BaseModel):
    temp_in: float = 25.0
    temp_out: float = 35.0
    pressure: float = 15.0
    vitesse: int = 1500


# ── Endpoint : Activer/Désactiver le mode manuel ─────
@app.post("/manual/on", summary="Activer le mode manuel")
def manual_on():
    """Active le mode manuel — les données viennent des sliders React."""
    global manual_mode
    manual_mode = True
    print("🎛️  Mode MANUEL activé")
    return {"status": "ok", "manual_mode": True}


@app.post("/manual/off", summary="Désactiver le mode manuel")
def manual_off():
    """Désactive le mode manuel — retour à simulation/Arduino."""
    global manual_mode
    manual_mode = False
    print("🎛️  Mode MANUEL désactivé — retour auto")
    return {"status": "ok", "manual_mode": False}


# ── Endpoint : Envoyer des valeurs manuelles ─────────
@app.post("/manual/set", summary="Envoyer des valeurs manuelles")
def manual_set(ctrl: ManualControl):
    """
    Reçoit les valeurs des sliders React et les stocke.
    Le thread de lecture va les utiliser pour générer le point de données.
    """
    global manual_mode
    manual_mode = True

    with manual_lock:
        manual_data["temp_in"] = ctrl.temp_in
        manual_data["temp_out"] = ctrl.temp_out
        manual_data["pressure"] = ctrl.pressure
        manual_data["vitesse"] = ctrl.vitesse

    return {
        "status": "ok",
        "manual_mode": True,
        "values": {
            "temp_in": ctrl.temp_in,
            "temp_out": ctrl.temp_out,
            "pressure": ctrl.pressure,
            "vitesse": ctrl.vitesse,
        },
    }


# ── Endpoint : Export Excel ───────────────────────────
@app.get("/export/excel", summary="Exporter l'historique en Excel (.xlsx)")
def export_excel():
    """
    Génère un fichier Excel à la volée.
    - Lignes ALARME → fond ROUGE
    - Lignes NORMALES → fond VERT
    - Lignes ATTENTION → fond ORANGE
    """
    if openpyxl is None:
        return {"error": "pip install openpyxl"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historique VFD"

    # En-têtes
    headers = [
        "Date / Heure",
        "Temp Entrée (°C)",
        "Temp Sortie (°C)",
        "Delta T (°C)",
        "Pression (Bar)",
        "Vitesse (RPM)",
        "Fréquence (Hz)",
        "Moteurs",
        "Mode",
        "Statut",
    ]
    ws.append(headers)

    # Styles pour les en-têtes
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Styles pour les lignes de données
    red_fill = PatternFill(start_color="FF4D4D", end_color="FF4D4D", fill_type="solid")
    green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
    orange_fill = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")

    # Copier l'historique
    hist_copy = list(history)

    for point in hist_copy:
        # Déterminer le statut
        if point.get("alarm"):
            statut = "ALARME"
        elif point.get("warning"):
            statut = "ATTENTION"
        else:
            statut = "NORMAL"

        row = [
            point.get("timestamp", "")[:19].replace("T", " "),
            point.get("temp_in", 0),
            point.get("temp_out", 0),
            point.get("delta_t", 0),
            point.get("pressure", 0),
            point.get("vitesse", 0),
            point.get("freq", 0),
            point.get("motors", 0),
            point.get("mode", ""),
            statut,
        ]
        ws.append(row)

        # Colorer la ligne
        row_num = ws.max_row
        if statut == "ALARME":
            for cell in ws[row_num]:
                cell.fill = red_fill
        elif statut == "ATTENTION":
            for cell in ws[row_num]:
                cell.fill = orange_fill
        else:
            for cell in ws[row_num]:
                cell.fill = green_fill

    # Ajuster la largeur des colonnes
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = max(len(header) + 4, 16)

    # Sauvegarder dans un buffer mémoire
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Historique_VFD_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers_dict = {
        "Content-Disposition": f"attachment; filename={filename}"
    }

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_dict,
    )


# ── Endpoint : Configuration ─────────────────────────
@app.get("/config", summary="Configuration actuelle")
def get_config() -> dict:
    with config_lock:
        return {
            **dict(config),
            "manual_mode": manual_mode,
        }


# ── Endpoint : Modifier les seuils ───────────────────
@app.post("/config/thresholds", summary="Modifier les seuils d'alarme")
def set_thresholds(
    alarm_temp: float = Query(default=None, ge=0, le=100),
    alarm_pression: float = Query(default=None, ge=0, le=50),
    warn_temp: float = Query(default=None, ge=0, le=100),
    warn_pression: float = Query(default=None, ge=0, le=50),
) -> dict:
    """Met à jour les seuils d'alarme/avertissement."""
    with config_lock:
        if alarm_temp is not None:
            config["alarm_temp"] = alarm_temp
        if alarm_pression is not None:
            config["alarm_pression"] = alarm_pression
        if warn_temp is not None:
            config["warn_temp"] = warn_temp
        if warn_pression is not None:
            config["warn_pression"] = warn_pression
        return {"status": "ok", "config": dict(config)}


# ── WebSocket Endpoint ────────────────────────────────
@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    """WebSocket temps réel."""
    await manager.connect(websocket)
    try:
        # Envoyer immédiatement la dernière mesure
        with data_lock:
            await websocket.send_json(dict(latest_data))

        # Garder la connexion ouverte
        while True:
            try:
                msg = await asyncio.wait_for(
                    websocket.receive_text(),
                    timeout=30,
                )
                if msg == "ping":
                    await websocket.send_text("pong")
            except asyncio.TimeoutError:
                try:
                    await websocket.send_json({"type": "heartbeat"})
                except Exception:
                    break

    except WebSocketDisconnect:
        pass
    except Exception as e:
        print(f"⚠️  WebSocket erreur: {e}")
    finally:
        await manager.disconnect(websocket)


# ─────────────────────────────────────────
#  DÉMARRAGE
# ─────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print("  ONCF Z2M — Supervision Refroidissement VFD")
    print("  Backend FastAPI v5.0")
    print("=" * 60)
    print(f"  Serial     : {SERIAL_PORT} @ {BAUD_RATE} baud")
    print(f"  API REST   : http://localhost:8000/data")
    print(f"  WebSocket  : ws://localhost:8000/ws")
    print(f"  Manuel     : POST http://localhost:8000/manual/set")
    print(f"  Excel      : http://localhost:8000/export/excel")
    print("=" * 60)

    # Démarrer le thread de lecture
    thread = threading.Thread(target=boucle_lecture, daemon=True)
    thread.start()

    # Démarrer serveur FastAPI
    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="warning")
